from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from workinghours import from_xls_to_json

# Constants
HEADER_FONT = Font(bold=True, size=14)
TEXT_FONT = Font(size=14)
ALIGN_CENTER = Alignment(horizontal="center")
ALIGN_LEFT = Alignment(horizontal="left")
ALIGN_RIGHT = Alignment(horizontal="right")
THICK_BORDER = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick"),
)

# File path constants
MONTHLY_ATTENDANCE_REPORT_XLS_PATH = "../input/report_example.xlsx"
WORKER_DETAILS_JSON_PATH = "../config/id2worker.json"
SALARY_DETAILS_OUTPUT_PATH = "../output/salary_details.xlsx"


def load_workers_monthly_attendance_data(
    monthly_attendance_report_xsl_path: str, worker_details_json_path: str
) -> List[Dict[str, Any]]:
    """Load workers data from Excel and JSON files."""
    try:
        return from_xls_to_json(
            monthly_attendance_report_xsl_path, worker_details_json_path
        )
    except Exception as e:
        print(f"Error loading workers data: {e}")
        return []


def create_workbook() -> openpyxl.Workbook:
    """Create a new workbook and set initial properties."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "Salary Details"
    return wb


def write_worker_data(
    ws: Worksheet,
    worker: Dict[str, Any],
    start_row: int,
) -> int:
    """Write worker data to the worksheet starting from the given row."""
    ws.cell(row=start_row, column=1, value=worker["name"]).font = HEADER_FONT
    ws.cell(row=start_row, column=1).alignment = ALIGN_RIGHT

    # Writing title in the second row of the table
    ws.cell(row=start_row + 1, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 1, column=2, value="שעות").font = HEADER_FONT
    ws.cell(row=start_row + 1, column=2).border = THICK_BORDER
    ws.cell(row=start_row + 1, column=3, value="לשעה ₪").font = HEADER_FONT
    ws.cell(row=start_row + 1, column=3).border = THICK_BORDER
    ws.cell(row=start_row + 1, column=4, value='סה"כ ₪').font = HEADER_FONT
    ws.cell(row=start_row + 1, column=4).border = THICK_BORDER

    # Writing regular hours row in the 3rd row of the table
    ws.cell(row=start_row + 2, column=1, value="ש.רגילות").font = HEADER_FONT
    ws.cell(row=start_row + 2, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 2, column=2, value=worker["hours"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 2, column=2).border = THICK_BORDER
    ws.cell(row=start_row + 2, column=3, value=worker["per_hour"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 2, column=3).border = THICK_BORDER
    ws.cell(
        row=start_row + 2, column=4, value=worker["reg_hours_sal"]
    ).font = TEXT_FONT
    ws.cell(row=start_row + 2, column=4).border = THICK_BORDER

    # Empty row (start_row + 3)

    # Writing extra hours row in the 4th row of the table
    ws.cell(row=start_row + 4, column=1, value="ש.נ. 125%").font = HEADER_FONT
    ws.cell(row=start_row + 4, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 4, column=2, value=worker["hours_125"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 4, column=2).border = THICK_BORDER
    ws.cell(row=start_row + 4, column=3, value=worker["per_hour_125"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 4, column=3).border = THICK_BORDER
    ws.cell(
        row=start_row + 4, column=4, value=worker["extra_hours_sal"]
    ).font = TEXT_FONT
    ws.cell(row=start_row + 4, column=4).border = THICK_BORDER

    # Empty row (start_row + 5)

    # Writing transportation expenses row in the 6th row of the table
    ws.cell(row=start_row + 6, column=1, value="נסיעות").font = HEADER_FONT
    ws.cell(row=start_row + 6, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 6, column=2).border = THICK_BORDER
    ws.cell(row=start_row + 6, column=3).border = THICK_BORDER
    ws.cell(
        row=start_row + 6, column=4, value=worker["trans_expanses"]
    ).font = TEXT_FONT
    ws.cell(row=start_row + 6, column=4).border = THICK_BORDER

    # Writing total salary row in the 7th row of the table
    ws.cell(row=start_row + 7, column=1, value='סה"כ').font = HEADER_FONT
    ws.cell(row=start_row + 7, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 7, column=2, value=worker["total_hours"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 7, column=2).border = THICK_BORDER
    ws.cell(row=start_row + 7, column=3).border = THICK_BORDER
    ws.cell(row=start_row + 7, column=4, value=worker["total_sal"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 7, column=4).border = THICK_BORDER

    # Empty row (start_row + 8)

    # Writing work days row in the 9th row of the table
    ws.cell(row=start_row + 9, column=1, value="ימי עבודה").font = HEADER_FONT
    ws.cell(row=start_row + 9, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 9, column=2, value=worker["work_days"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 9, column=2).border = THICK_BORDER

    # Writing holiday row in the 10th row of the table
    ws.cell(row=start_row + 10, column=1, value="חג").font = HEADER_FONT
    ws.cell(row=start_row + 10, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 10, column=2, value=worker["holidays"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 10, column=2).border = THICK_BORDER

    # Writing gift row in the 11th row of the table
    ws.cell(row=start_row + 11, column=1, value="מתנה").font = HEADER_FONT
    ws.cell(row=start_row + 11, column=1).border = THICK_BORDER
    ws.cell(
        row=start_row + 11, column=2, value=worker["holiday_present"]
    ).font = TEXT_FONT
    ws.cell(row=start_row + 11, column=2).border = THICK_BORDER

    # Writing sick days row in the 12th row of the table
    ws.cell(row=start_row + 12, column=1, value="ימי מחלה").font = HEADER_FONT
    ws.cell(row=start_row + 12, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 12, column=2, value=worker["sick_days"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 12, column=2).border = THICK_BORDER

    # Writing vacation days row in the 13th row of the table
    ws.cell(row=start_row + 13, column=1, value="ימי חופש").font = HEADER_FONT
    ws.cell(row=start_row + 13, column=1).border = THICK_BORDER
    ws.cell(row=start_row + 13, column=2, value=worker["vac_days"]).font = (
        TEXT_FONT
    )
    ws.cell(row=start_row + 13, column=2).border = THICK_BORDER

    # Write absense hours row in the 14th row of the table
    ws.cell(row=start_row + 14, column=1, value="שעות להוריד").font = (
        HEADER_FONT
    )
    ws.cell(row=start_row + 14, column=1).border = THICK_BORDER
    ws.cell(
        row=start_row + 14, column=2, value=worker["absense_hours"]
    ).font = TEXT_FONT
    ws.cell(row=start_row + 14, column=2).border = THICK_BORDER

    # Adding some space between tables
    return start_row + 18


def set_column_widths(ws: Worksheet) -> None:
    """Set column widths for better readability."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def save_workbook(wb: openpyxl.Workbook, file_path: str) -> None:
    """Save the workbook to the specified file path."""
    try:
        wb.save(file_path)
    except Exception as e:
        print(f"Error saving workbook: {e}")


def main() -> None:
    workers = load_workers_monthly_attendance_data(
        MONTHLY_ATTENDANCE_REPORT_XLS_PATH, WORKER_DETAILS_JSON_PATH
    )
    wb = create_workbook()
    ws = wb.active

    start_row = 1
    for worker in workers:
        start_row = write_worker_data(ws, worker, start_row)

    set_column_widths(ws)
    save_workbook(wb, SALARY_DETAILS_OUTPUT_PATH)


if __name__ == "__main__":
    main()
